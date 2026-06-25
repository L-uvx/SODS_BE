import zipfile
from pathlib import Path
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.application.polygon_obstacle_excel_parser import (
    PolygonObstacleExcelParseError,
    parse_polygon_obstacle_excel,
    _parse_dms,
    _parse_dms_components,
    _detect_obstacle_template,
)


def _build_workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    for row in rows:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class TestParseDmsComponents:
    def test_all_components_provided(self) -> None:
        result = _parse_dms_components(103.0, 58.0, 33.11)
        assert result == pytest.approx(103.975864, rel=1e-6)

    def test_minutes_and_seconds_zero(self) -> None:
        result = _parse_dms_components(30.0, 0.0, 0.0)
        assert result == 30.0

    def test_seconds_as_none(self) -> None:
        result = _parse_dms_components(103.0, 58.0, None)
        assert result == pytest.approx(103.966667, rel=1e-6)

    def test_minutes_as_none(self) -> None:
        result = _parse_dms_components(103.0, None, 30.0)
        assert result == pytest.approx(103.008333, rel=1e-6)

    def test_all_none(self) -> None:
        result = _parse_dms_components(None, None, None)
        assert result == 0.0

    def test_degrees_as_none(self) -> None:
        result = _parse_dms_components(None, 30.0, 0.0)
        assert result == 0.5


class TestParseDms:
    def test_standard_format(self) -> None:
        result = _parse_dms("103°58'33.11\"", field_name="lon", row_number=2)
        assert result == pytest.approx(103.975864, rel=1e-6)

    def test_curly_double_quote_seconds(self) -> None:
        result = _parse_dms("103°58'33.11\u201d", field_name="lon", row_number=2)
        assert result == pytest.approx(103.975864, rel=1e-6)

    def test_curly_single_quote_minutes(self) -> None:
        result = _parse_dms("103°58\u201933.11\"", field_name="lon", row_number=2)
        assert result == pytest.approx(103.975864, rel=1e-6)

    def test_fullwidth_quote_seconds(self) -> None:
        result = _parse_dms("103°58'33.11\uff02", field_name="lon", row_number=2)
        assert result == pytest.approx(103.975864, rel=1e-6)


class TestDetectObstacleTemplate:
    def test_detects_8col_when_8_columns(self) -> None:
        rows = [["名称", "经度度", "经度分", "经度秒", "纬度度", "纬度分", "纬度秒", "高程"]]
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Sheet1"
        for r in rows:
            ws.append(r)
        template = _detect_obstacle_template(ws)
        assert template == "8col"

    def test_detects_8col_when_more_than_8_columns(self) -> None:
        rows = [["名称", "a", "b", "c", "d", "e", "f", "g", "h", "i"]]
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Sheet1"
        for r in rows:
            ws.append(r)
        template = _detect_obstacle_template(ws)
        assert template == "8col"

    def test_detects_4col_when_4_columns(self) -> None:
        rows = [["名称", "经度", "纬度", "高程"]]
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Sheet1"
        for r in rows:
            ws.append(r)
        template = _detect_obstacle_template(ws)
        assert template == "4col"


class TestParse8ColTemplate:
    def test_parses_single_obstacle_with_8col(self) -> None:
        rows = [
            ["障碍物名称", "经度度", "经度分", "经度秒", "纬度度", "纬度分", "纬度秒", "顶部高程"],
            ["障碍物1", 103, 58, 33.11, 30, 30, 24.77, 549.9],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        obstacles = parse_polygon_obstacle_excel(wb_bytes)
        assert len(obstacles) == 1
        obs = obstacles[0]
        assert obs.name == "障碍物1"
        assert len(obs.points) == 1
        assert obs.points[0].longitude_decimal == pytest.approx(103.975864, rel=1e-6)
        assert obs.points[0].latitude_decimal == pytest.approx(30.506881, rel=1e-6)
        assert obs.top_elevation == 549.9

    def test_parses_multiple_rows_grouped_by_name_8col(self) -> None:
        rows = [
            ["名称", "d", "m", "s", "d", "m", "s", "elev"],
            ["障碍物A", 103, 58, 33.0, 30, 30, 0.0, 100.0],
            ["障碍物A", 103, 59, 0.0, 30, 31, 0.0, 100.0],
            ["障碍物B", 104, 0, 0.0, 31, 0, 0.0, 200.0],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        obstacles = parse_polygon_obstacle_excel(wb_bytes)
        assert len(obstacles) == 2
        assert obstacles[0].name == "障碍物A"
        assert len(obstacles[0].points) == 2
        assert obstacles[1].name == "障碍物B"
        assert len(obstacles[1].points) == 1

    def test_8col_sets_longitude_text_and_latitude_text_to_empty(self) -> None:
        rows = [
            ["名称", "d", "m", "s", "d", "m", "s", "elev"],
            ["障碍物1", 103, 58, 33.0, 30, 30, 0.0, 100.0],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        obstacles = parse_polygon_obstacle_excel(wb_bytes)
        assert obstacles[0].points[0].longitude_text == ""
        assert obstacles[0].points[0].latitude_text == ""

    def test_8col_treats_none_components_as_zero(self) -> None:
        rows = [
            ["名称", "d", "m", "s", "d", "m", "s", "elev"],
            ["障碍物1", None, 30, None, 0, 0, None, 100.0],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        obstacles = parse_polygon_obstacle_excel(wb_bytes)
        assert obstacles[0].points[0].longitude_decimal == pytest.approx(0.5, rel=1e-6)
        assert obstacles[0].points[0].latitude_decimal == 0.0


class TestBackwardCompat4Col:
    def test_parse_fixed_template_groups_rows_into_obstacles(self) -> None:
        excel_bytes = Path("docs/import_demo.xlsx").read_bytes()
        obstacles = parse_polygon_obstacle_excel(excel_bytes)
        assert len(obstacles) == 2
        first_obstacle = obstacles[0]
        assert first_obstacle.name == "障碍物1"
        assert len(first_obstacle.points) == 5
        assert first_obstacle.top_elevation == 549.9
        assert [point.row_number for point in first_obstacle.points] == [2, 3, 4, 5, 6]
        assert round(first_obstacle.points[0].longitude_decimal, 6) == 103.975864
        assert round(first_obstacle.points[0].latitude_decimal, 6) == 30.506881
        second_obstacle = obstacles[1]
        assert second_obstacle.name == "障碍物2"
        assert len(second_obstacle.points) == 5
        assert second_obstacle.top_elevation == 549.9
        assert [point.row_number for point in second_obstacle.points] == [7, 8, 9, 10, 11]

    def test_parses_4col_simple(self) -> None:
        rows = [
            ["障碍物名称", "经度", "纬度", "顶部高程"],
            ["障碍物1", "103°58'33.11\"", "030°30'24.77\"", 549.9],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        obstacles = parse_polygon_obstacle_excel(wb_bytes)
        assert len(obstacles) == 1
        assert obstacles[0].name == "障碍物1"
        assert obstacles[0].points[0].longitude_decimal == pytest.approx(103.975864, rel=1e-6)

    def test_4col_longitude_text_is_preserved(self) -> None:
        rows = [
            ["x", "y", "z", "w"],
            ["测试物", "103°58'33.11\"", "030°30'24.77\"", 100.0],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        obstacles = parse_polygon_obstacle_excel(wb_bytes)
        assert obstacles[0].points[0].longitude_text == "103°58'33.11\""
        assert obstacles[0].points[0].latitude_text == "030°30'24.77\""


class TestHeaderValidationRemoved:
    def test_any_row1_content_is_accepted_4col(self) -> None:
        rows = [
            ["anything", "goes", "here", "now"],
            ["障碍物1", "103°58'33.11\"", "030°30'24.77\"", 549.9],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        obstacles = parse_polygon_obstacle_excel(wb_bytes)
        assert len(obstacles) == 1

    def test_any_row1_content_is_accepted_8col(self) -> None:
        rows = [
            [1, 2, 3, 4, 5, 6, 7, 8],
            ["障碍物1", 103, 58, 33.0, 30, 30, 0.0, 100.0],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        obstacles = parse_polygon_obstacle_excel(wb_bytes)
        assert len(obstacles) == 1


class TestIndexBasedSheet:
    def test_no_worksheets_raises_error(self) -> None:
        """Empty workbook (no sheets) should raise PolygonObstacleExcelParseError."""
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(
                "[Content_Types].xml",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                '<Default Extension="xml" ContentType="application/xml"/>'
                '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
                "</Types>",
            )
            zf.writestr(
                "_rels/.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
                "</Relationships>",
            )
            zf.writestr(
                "xl/workbook.xml",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                "<sheets/>"
                "</workbook>",
            )
            zf.writestr(
                "xl/_rels/workbook.xml.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>',
            )
        with pytest.raises(PolygonObstacleExcelParseError, match="工作簿中没有工作表"):
            parse_polygon_obstacle_excel(buf.getvalue())

    def test_uses_first_sheet_regardless_of_name(self) -> None:
        """Sheet with a non-standard name at index 0 should still be parsed."""
        wb = Workbook()
        ws = wb.active
        ws.title = "CustomName"
        ws.cell(row=2, column=1, value="障碍物A")
        ws.cell(row=2, column=2, value="103°30'0\"")
        ws.cell(row=2, column=3, value="30°30'0\"")
        ws.cell(row=2, column=4, value=100)
        buf = BytesIO()
        wb.save(buf)
        result = parse_polygon_obstacle_excel(buf.getvalue())
        assert len(result) == 1
        assert result[0].name == "障碍物A"

    def test_ignores_second_sheet(self) -> None:
        """Only the first sheet (index 0) should be parsed; second sheet ignored."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "First"
        ws1.cell(row=2, column=1, value="障碍物A")
        ws1.cell(row=2, column=2, value="103°30'0\"")
        ws1.cell(row=2, column=3, value="30°30'0\"")
        ws1.cell(row=2, column=4, value=100)
        ws2 = wb.create_sheet("Second")
        ws2.cell(row=2, column=1, value="障碍物B")
        ws2.cell(row=2, column=2, value="104°30'0\"")
        ws2.cell(row=2, column=3, value="31°30'0\"")
        ws2.cell(row=2, column=4, value=200)
        buf = BytesIO()
        wb.save(buf)
        result = parse_polygon_obstacle_excel(buf.getvalue())
        assert len(result) == 1
        assert result[0].name == "障碍物A"


class TestErrorCases:
    def test_raises_for_inconsistent_top_elevation_with_same_name(self) -> None:
        workbook_bytes = _build_workbook_bytes(
            [
                ["障碍物名称", "经度", "纬度", "顶部高程"],
                ["障碍物1", "103°58'33.11\"", "030°30'24.77\"", 549.9],
                ["障碍物1", "103°58'41.20\"", "030°30'20.34\"", 550.1],
            ]
        )
        with pytest.raises(PolygonObstacleExcelParseError) as exc_info:
            parse_polygon_obstacle_excel(workbook_bytes)
        assert "顶部高程" in str(exc_info.value)

    def test_raises_for_invalid_excel_file_bytes(self) -> None:
        with pytest.raises(PolygonObstacleExcelParseError) as exc_info:
            parse_polygon_obstacle_excel(b"not-an-excel-file")
        assert str(exc_info.value) == "invalid excel file"

    def test_raises_for_blank_name(self) -> None:
        rows = [
            ["x", "y", "z", "w"],
            ["", "103°58'33.11\"", "030°30'24.77\"", 100.0],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        with pytest.raises(PolygonObstacleExcelParseError) as exc_info:
            parse_polygon_obstacle_excel(wb_bytes)
        assert "障碍物名称" in str(exc_info.value)

    def test_raises_for_missing_lon_8col(self) -> None:
        rows = [
            ["a", "b", "c", "d", "e", "f", "g", "h"],
            ["障碍物1", None, None, None, 30, 30, 0.0, 100.0],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        with pytest.raises(PolygonObstacleExcelParseError) as exc_info:
            parse_polygon_obstacle_excel(wb_bytes)
        assert "经度" in str(exc_info.value)

    def test_raises_for_missing_lat_8col(self) -> None:
        rows = [
            ["a", "b", "c", "d", "e", "f", "g", "h"],
            ["障碍物1", 103.0, 58.0, 33.0, None, None, None, 100.0],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        with pytest.raises(PolygonObstacleExcelParseError) as exc_info:
            parse_polygon_obstacle_excel(wb_bytes)
        assert "纬度" in str(exc_info.value)

    def test_skips_blank_rows(self) -> None:
        rows = [
            ["a", "b", "c", "d"],
            [None, None, None, None],
            ["", "", "", ""],
            ["障碍物1", "103°58'33.11\"", "030°30'24.77\"", 100.0],
        ]
        wb_bytes = _build_workbook_bytes(rows)
        obstacles = parse_polygon_obstacle_excel(wb_bytes)
        assert len(obstacles) == 1

    def test_raises_for_empty_worksheet(self) -> None:
        rows = [["a", "b", "c", "d"]]
        wb_bytes = _build_workbook_bytes(rows)
        with pytest.raises(PolygonObstacleExcelParseError) as exc_info:
            parse_polygon_obstacle_excel(wb_bytes)
        assert "no data rows" in str(exc_info.value)
