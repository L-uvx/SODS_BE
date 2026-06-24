import zipfile
from io import BytesIO

from openpyxl import Workbook
import pytest

from app.application.point_obstacle_excel_parser import (
    PointObstacleExcelParseError,
    parse_point_obstacle_excel,
)


def _build_workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    for row in rows:
        worksheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parse_point_obstacle_excel_deduplicates_identical_rows() -> None:
    workbook_bytes = _build_workbook_bytes(
        [
            ["障碍物名称", "经度", "纬度", "顶部高程"],
            ["点障碍物1", "103°58'33.11\"", "030°30'24.77\"", 549.9],
            ["点障碍物1", "103°58'33.11\"", "030°30'24.77\"", 549.9],
            ["点障碍物2", "103°58'41.20\"", "030°30'20.34\"", 550.1],
        ]
    )

    obstacles = parse_point_obstacle_excel(workbook_bytes)

    assert len(obstacles) == 2
    assert obstacles[0].name == "点障碍物1"
    assert obstacles[0].top_elevation == 549.9
    assert obstacles[0].row_number == 2
    assert round(obstacles[0].longitude_decimal, 6) == 103.975864
    assert round(obstacles[0].latitude_decimal, 6) == 30.506881
    assert obstacles[1].name == "点障碍物2"
    assert obstacles[1].row_number == 4


def test_parse_point_obstacle_excel_raises_for_same_name_with_different_coordinates() -> None:
    workbook_bytes = _build_workbook_bytes(
        [
            ["障碍物名称", "经度", "纬度", "顶部高程"],
            ["点障碍物1", "103°58'33.11\"", "030°30'24.77\"", 549.9],
            ["点障碍物1", "103°58'41.20\"", "030°30'24.77\"", 549.9],
        ]
    )

    with pytest.raises(PointObstacleExcelParseError) as exc_info:
        parse_point_obstacle_excel(workbook_bytes)

    assert "坐标或顶部高程" in str(exc_info.value)


def test_parse_point_obstacle_excel_raises_for_same_name_with_different_top_elevation() -> None:
    workbook_bytes = _build_workbook_bytes(
        [
            ["障碍物名称", "经度", "纬度", "顶部高程"],
            ["点障碍物1", "103°58'33.11\"", "030°30'24.77\"", 549.9],
            ["点障碍物1", "103°58'33.11\"", "030°30'24.77\"", 550.1],
        ]
    )

    with pytest.raises(PointObstacleExcelParseError) as exc_info:
        parse_point_obstacle_excel(workbook_bytes)

    assert "坐标或顶部高程" in str(exc_info.value)


def test_parse_point_obstacle_excel_raises_for_invalid_excel_file_bytes() -> None:
    with pytest.raises(PointObstacleExcelParseError) as exc_info:
        parse_point_obstacle_excel(b"not-an-excel-file")

    assert str(exc_info.value) == "invalid excel file"


def test_parse_point_obstacle_excel_8col_template() -> None:
    workbook_bytes = _build_workbook_bytes(
        [
            ["障碍物名称", "经度_度", "经度_分", "经度_秒", "纬度_度", "纬度_分", "纬度_秒", "顶部高程"],
            ["点障碍物1", 103, 58, 33.11, 30, 30, 24.77, 549.9],
        ]
    )

    obstacles = parse_point_obstacle_excel(workbook_bytes)

    assert len(obstacles) == 1
    assert obstacles[0].name == "点障碍物1"
    assert obstacles[0].longitude_text == ""
    assert obstacles[0].latitude_text == ""
    assert round(obstacles[0].longitude_decimal, 6) == 103.975864
    assert round(obstacles[0].latitude_decimal, 6) == 30.506881
    assert obstacles[0].top_elevation == 549.9


def test_parse_point_obstacle_excel_4col_template_backward_compat() -> None:
    workbook_bytes = _build_workbook_bytes(
        [
            ["障碍物名称", "经度", "纬度", "顶部高程"],
            ["点障碍物1", "103°58'33.11\"", "030°30'24.77\"", 549.9],
        ]
    )

    obstacles = parse_point_obstacle_excel(workbook_bytes)

    assert len(obstacles) == 1
    assert obstacles[0].name == "点障碍物1"
    assert obstacles[0].longitude_text == "103°58'33.11\""
    assert obstacles[0].latitude_text == "030°30'24.77\""
    assert round(obstacles[0].longitude_decimal, 6) == 103.975864
    assert round(obstacles[0].latitude_decimal, 6) == 30.506881


def test_parse_point_obstacle_excel_8col_deduplication() -> None:
    workbook_bytes = _build_workbook_bytes(
        [
            ["障碍物名称", "经度_度", "经度_分", "经度_秒", "纬度_度", "纬度_分", "纬度_秒", "顶部高程"],
            ["点障碍物1", 103, 58, 33.11, 30, 30, 24.77, 549.9],
            ["点障碍物1", 103, 58, 33.11, 30, 30, 24.77, 549.9],
            ["点障碍物2", 103, 58, 41.20, 30, 30, 20.34, 550.1],
        ]
    )

    obstacles = parse_point_obstacle_excel(workbook_bytes)

    assert len(obstacles) == 2


def test_parse_point_obstacle_excel_8col_with_none_components() -> None:
    workbook_bytes = _build_workbook_bytes(
        [
            ["障碍物名称", "经度_度", "经度_分", "经度_秒", "纬度_度", "纬度_分", "纬度_秒", "顶部高程"],
            ["点障碍物1", 103, None, None, 30, None, None, 549.9],
        ]
    )

    obstacles = parse_point_obstacle_excel(workbook_bytes)

    assert len(obstacles) == 1
    assert round(obstacles[0].longitude_decimal, 1) == 103.0
    assert round(obstacles[0].latitude_decimal, 1) == 30.0
    assert obstacles[0].longitude_text == ""
    assert obstacles[0].latitude_text == ""


def test_parse_point_obstacle_excel_no_header_validation() -> None:
    workbook_bytes = _build_workbook_bytes(
        [
            ["名称", "Longitude", "Latitude", "Elevation"],
            ["点障碍物1", "103°58'33.11\"", "030°30'24.77\"", 549.9],
        ]
    )

    obstacles = parse_point_obstacle_excel(workbook_bytes)

    assert len(obstacles) == 1
    assert obstacles[0].name == "点障碍物1"
    assert round(obstacles[0].longitude_decimal, 6) == 103.975864


def test_no_worksheets_raises_error() -> None:
    """Empty workbook (no sheets) should raise PointObstacleExcelParseError."""
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Override PartName="/xl/workbook.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            "</Types>",
        )
        zf.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
            'Target="xl/workbook.xml"/>'
            "</Relationships>",
        )
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            "<sheets/>"
            "</workbook>",
        )
    with pytest.raises(PointObstacleExcelParseError, match="工作簿中没有工作表"):
        parse_point_obstacle_excel(buf.getvalue())


def test_uses_first_sheet_regardless_of_name() -> None:
    """Sheet with a non-standard name at index 0 should still be parsed."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CustomName"
    ws.cell(row=2, column=1, value="障碍物A")
    ws.cell(row=2, column=2, value="103°30'0\"")
    ws.cell(row=2, column=3, value="30°30'0\"")
    ws.cell(row=2, column=4, value=100)
    buf = BytesIO()
    wb.save(buf)
    result = parse_point_obstacle_excel(buf.getvalue())
    assert len(result) == 1
    assert result[0].name == "障碍物A"


def test_ignores_second_sheet() -> None:
    """Only the first sheet (index 0) should be parsed; second sheet ignored."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1.cell(row=2, column=1, value="障碍物A")
    ws1.cell(row=2, column=2, value="103°30'0\"")
    ws1.cell(row=2, column=3, value="30°30'0\"")
    ws1.cell(row=2, column=4, value=100)
    ws2 = wb.create_sheet("Second")
    ws2.cell(row=2, column=1, value="障碍物B")
    ws2.cell(row=2, column=2, value="104°30'0\"")
    ws2.cell(row=2, column=3, value="31°30'0\"")
    ws2.cell(row=2, column=4, value=200)
    buf = BytesIO()
    wb.save(buf)
    result = parse_point_obstacle_excel(buf.getvalue())
    assert len(result) == 1
    assert result[0].name == "障碍物A"
