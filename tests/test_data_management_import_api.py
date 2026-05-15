import pytest
from io import BytesIO
from pathlib import Path
from collections.abc import Generator
from contextlib import contextmanager

from openpyxl import Workbook
from pydantic import ValidationError
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool
from fastapi.testclient import TestClient

from app.db.base import Base
from app.models.airport import Airport
from app.models.runway import Runway
from app.models.station import Station
from app.main import app
from app.api.deps import get_db_session
from app.application.data_management_import import (
    _parse_degree,
    _get_number_from_string,
    _extract_runway_info,
    _int_floor,
    _open_workbook,
    _parse_runway_sheet,
    _parse_station_sheet,
    _import_airport_from_excel,
    AirportImportParseError,
)
from app.schemas.data_management import (
    AirportImportResponse,
    AirportUpsertRequest,
    RunwayUpsertRequest,
    StationUpsertRequest,
)


# --- Fixtures ---

@pytest.fixture(scope="module")
def dualiang_xlsx_bytes() -> bytes:
    path = Path(__file__).parent.parent / "docs" / "temp" / "XX机场.xlsx"
    return path.read_bytes()


@pytest.fixture(scope="module")
def empty_airport_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws1 = wb.create_sheet("Sheet1", 0)
    for i, h in enumerate([
        "跑道号码", "跑道真方位（°）", "跑道长度（米）", "跑道宽度（米）",
        "跑道入口标高（米）", "最大适航机型（H：航空器高度）",
        "跑道中心点经度", "跑道中心点纬度", "机场基准点标高（米）",
        "跑道类型", "编码A", "编码B", "仪表着陆系统类别",
        "最大可起降航空器类别类别",
    ], start=1):
        ws1.cell(row=1, column=i, value=h)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@contextmanager
def _create_import_test_session() -> Generator[Session, None, None]:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    testing_session_local = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(
        bind=engine,
        tables=[Airport.__table__, Runway.__table__, Station.__table__],
    )
    session = testing_session_local()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(
            bind=engine,
            tables=[Station.__table__, Runway.__table__, Airport.__table__],
        )


# --- Utility Tests ---

class TestParseDegree:
    def test_decimal_degrees_returns_float(self) -> None:
        assert _parse_degree("103.94371") == pytest.approx(103.94371)

    def test_dms_with_degree_symbol(self) -> None:
        expected = 103.0 + 56.0 / 60.0 + 37.8 / 3600.0
        assert _parse_degree("103°56'37.8\"") == pytest.approx(expected)

    def test_dms_with_chinese_chars(self) -> None:
        expected = 103.0 + 56.0 / 60.0 + 37.8 / 3600.0
        assert _parse_degree("103度56分37.8秒") == pytest.approx(expected)

    def test_plain_float_string(self) -> None:
        assert _parse_degree("30.53463") == pytest.approx(30.53463)

    def test_none_returns_none(self) -> None:
        assert _parse_degree(None) is None

    def test_int_returns_float(self) -> None:
        assert _parse_degree(103) == pytest.approx(103.0)

    def test_float_returns_float(self) -> None:
        assert _parse_degree(103.5) == pytest.approx(103.5)

    def test_minutes_out_of_range_raises(self) -> None:
        with pytest.raises(ValueError):
            _parse_degree("103°65'37.8\"")

    def test_seconds_out_of_range_raises(self) -> None:
        with pytest.raises(ValueError):
            _parse_degree("103°56'65.0\"")

    def test_abbreviated_dms_degrees_only(self) -> None:
        assert _parse_degree("103°") == pytest.approx(103.0)

    def test_abbreviated_dms_degrees_minutes(self) -> None:
        assert _parse_degree("103°56'") == pytest.approx(103.0 + 56.0 / 60.0)

    def test_empty_string_returns_none(self) -> None:
        assert _parse_degree("") is None

    def test_whitespace_string_returns_none(self) -> None:
        assert _parse_degree("   ") is None

    def test_boolean_raises(self) -> None:
        with pytest.raises(ValueError):
            _parse_degree(True)


class TestGetNumberFromString:
    def test_plain_number(self) -> None:
        assert _get_number_from_string("3600") == pytest.approx(3600.0)

    def test_number_with_space(self) -> None:
        assert _get_number_from_string(" 3600 ") == pytest.approx(3600.0)

    def test_km_conversion(self) -> None:
        assert _get_number_from_string("3km") == pytest.approx(3000.0)

    def test_km_chinese(self) -> None:
        assert _get_number_from_string("3千米") == pytest.approx(3000.0)

    def test_nm_conversion(self) -> None:
        assert _get_number_from_string("1nm") == pytest.approx(1852.0)

    def test_float_with_unit(self) -> None:
        assert _get_number_from_string("3.5km") == pytest.approx(3500.0)

    def test_non_numeric_returns_none(self) -> None:
        # C# 对齐：无数字时返回 None，不抛异常
        assert _get_number_from_string("abc") is None

    def test_none_returns_none(self) -> None:
        assert _get_number_from_string(None) is None

    def test_integer_passed(self) -> None:
        assert _get_number_from_string(3600) == pytest.approx(3600.0)

    def test_float_passed(self) -> None:
        assert _get_number_from_string(3600.5) == pytest.approx(3600.5)

    def test_boolean_raises(self) -> None:
        with pytest.raises(ValueError):
            _get_number_from_string(True)


class TestExtractRunwayInfo:
    def test_typical_loc_name(self) -> None:
        assert _extract_runway_info("LOC20R") == "20R"

    def test_typical_gp_name(self) -> None:
        assert _extract_runway_info("GP02L") == "02L"

    def test_single_digit_runway(self) -> None:
        assert _extract_runway_info("LOC9L") == "9L"

    def test_fallback_last_word(self) -> None:
        assert _extract_runway_info("双流机场 02L") == "02L"

    def test_no_runway_number_returns_last_word(self) -> None:
        assert _extract_runway_info("西南近无方向信标台") == "西南近无方向信标台"

    def test_none_returns_none(self) -> None:
        assert _extract_runway_info(None) is None

    def test_empty_string_returns_none(self) -> None:
        assert _extract_runway_info("") is None

    def test_whitespace_string_returns_none(self) -> None:
        assert _extract_runway_info("   ") is None


class TestIntFloor:
    def test_positive(self) -> None:
        assert _int_floor(3.14159) == pytest.approx(3.14)

    def test_exact_two_decimals(self) -> None:
        assert _int_floor(3.14) == pytest.approx(3.14)

    def test_integer(self) -> None:
        assert _int_floor(3.0) == pytest.approx(3.0)

    def test_negative(self) -> None:
        assert _int_floor(-3.14159) == pytest.approx(-3.15)

    def test_zero(self) -> None:
        assert _int_floor(0.0) == pytest.approx(0.0)

    def test_large_value(self) -> None:
        assert _int_floor(123456.789) == pytest.approx(123456.78)


# --- Schema Tests ---

class TestAirportImportResponse:
    def test_minimal_response(self) -> None:
        resp = AirportImportResponse(id=1, airportName="Test", runwayCount=0, stationCount=0, warnings=[])
        data = resp.model_dump(by_alias=True)
        assert data == {"id": 1, "airportName": "Test", "runwayCount": 0, "stationCount": 0, "warnings": []}

    def test_with_warnings(self) -> None:
        resp = AirportImportResponse(
            id=1, airportName="Test", runwayCount=2, stationCount=3,
            warnings=[{"code": "w1", "message": "test warning"}],
        )
        data = resp.model_dump(by_alias=True)
        assert data["warnings"] == [{"code": "w1", "message": "test warning"}]


class TestDmsCompatibleValidators:
    def test_airport_longitude_accepts_dms_string(self) -> None:
        req = AirportUpsertRequest(name="Test", longitude="103°56'37.8\"", latitude="30.534")
        expected = 103.0 + 56.0 / 60.0 + 37.8 / 3600.0
        assert req.longitude == pytest.approx(expected)

    def test_airport_longitude_accepts_decimal_string(self) -> None:
        req = AirportUpsertRequest(name="Test", longitude="103.94", latitude="30.53")
        assert req.longitude == pytest.approx(103.94)

    def test_airport_longitude_accepts_float(self) -> None:
        req = AirportUpsertRequest(name="Test", longitude=103.94, latitude=30.53)
        assert req.longitude == pytest.approx(103.94)

    def test_airport_longitude_none(self) -> None:
        req = AirportUpsertRequest(name="Test", longitude=None, latitude=None)
        assert req.longitude is None

    def test_runway_longitude_accepts_dms(self) -> None:
        req = RunwayUpsertRequest(
            airportId=1, name="RWY", runNumber="18",
            longitude="103°56'37.8\"", latitude="30°32'4.5\"",
        )
        expected = 103.0 + 56.0 / 60.0 + 37.8 / 3600.0
        assert req.longitude == pytest.approx(expected)

    def test_station_longitude_accepts_dms(self) -> None:
        req = StationUpsertRequest(
            airportId=1, name="STN", stationType="NDB",
            longitude="103°56'37.8\"", latitude="30°32'4.5\"",
        )
        expected = 103.0 + 56.0 / 60.0 + 37.8 / 3600.0
        assert req.longitude == pytest.approx(expected)

    def test_airport_longitude_out_of_range_rejected(self) -> None:
        with pytest.raises(ValidationError):
            AirportUpsertRequest(name="Test", longitude=200, latitude=30.5)

    def test_airport_latitude_out_of_range_rejected(self) -> None:
        with pytest.raises(ValidationError):
            AirportUpsertRequest(name="Test", longitude=103.5, latitude=95.0)


# --- Excel Parser Tests ---

class TestOpenWorkbook:
    def test_valid_xlsx(self, dualiang_xlsx_bytes: bytes) -> None:
        wb = _open_workbook(dualiang_xlsx_bytes)
        assert "Sheet1" in wb.sheetnames

    def test_invalid_file_raises(self) -> None:
        with pytest.raises(AirportImportParseError):
            _open_workbook(b"not an excel file")


class TestParseRunwaySheet:
    def test_parses_real_file(self, dualiang_xlsx_bytes: bytes) -> None:
        rows = _parse_runway_sheet(dualiang_xlsx_bytes)
        assert len(rows) == 4

    def test_runway_fields_match(self, dualiang_xlsx_bytes: bytes) -> None:
        rows = _parse_runway_sheet(dualiang_xlsx_bytes)
        r = rows[0]
        assert r["run_number"] == "20L"
        assert r["name"] == "20L"
        assert r["direction"] == pytest.approx(202.0)
        assert r["length"] == pytest.approx(3600.0)
        assert r["width"] == pytest.approx(60.0)
        assert r["enter_height"] == pytest.approx(496.6)
        assert r["maximum_airworthiness"] == pytest.approx(3.0)
        assert r["longitude"] == pytest.approx(103.943710327148)
        assert r["latitude"] == pytest.approx(30.5346355438232)
        assert r["altitude"] == pytest.approx(496.6)
        assert r["runway_type"] == "精密进近跑道"
        assert r["runway_code_a"] == "4"
        assert r["runway_code_b"] == "F"
        assert r["station_sub_type"] == "I"

    def test_last_runway_fields(self, dualiang_xlsx_bytes: bytes) -> None:
        rows = _parse_runway_sheet(dualiang_xlsx_bytes)
        r = rows[3]
        assert r["run_number"] == "02L"
        assert r["station_sub_type"] == "II"

    def test_empty_sheet_returns_empty_list(self, empty_airport_xlsx_bytes: bytes) -> None:
        rows = _parse_runway_sheet(empty_airport_xlsx_bytes)
        assert rows == []

    def test_missing_sheet_raises(self) -> None:
        wb = Workbook()
        wb.create_sheet("OtherSheet", 0)
        buf = BytesIO()
        wb.save(buf)
        with pytest.raises(AirportImportParseError):
            _parse_runway_sheet(buf.getvalue())


class TestParseStationSheet:
    def test_parses_all_stations(self, dualiang_xlsx_bytes: bytes) -> None:
        rows = _parse_station_sheet(dualiang_xlsx_bytes)
        assert len(rows) == 21

    def test_loc_station_defaults(self, dualiang_xlsx_bytes: bytes) -> None:
        rows = _parse_station_sheet(dualiang_xlsx_bytes)
        locs = [r for r in rows if r["station_type"] == "LOC"]
        assert len(locs) > 0
        loc = locs[0]
        assert loc["coverage_radius"] == pytest.approx(46300.0)
        assert loc["antenna_hag"] is not None
        assert "fly_height" in loc
        assert loc["unit_number"] is not None
        assert loc["distance_endo_runway"] is not None

    def test_gp_station_defaults(self, dualiang_xlsx_bytes: bytes) -> None:
        rows = _parse_station_sheet(dualiang_xlsx_bytes)
        gps = [r for r in rows if r["station_type"] == "GP"]
        assert len(gps) > 0
        gp = gps[0]
        assert gp["coverage_radius"] == pytest.approx(18520.0)
        assert gp["antenna_hag"] is not None
        assert gp["downward_angle"] is not None
        assert gp["distance_to_runway"] is not None
        assert gp["distance_v_to_runway"] is not None

    def test_surface_detection_radar_has_runway_no(self, dualiang_xlsx_bytes: bytes) -> None:
        rows = _parse_station_sheet(dualiang_xlsx_bytes)
        sdrs = [r for r in rows if r["station_type"] == "Surface_Detection_Radar"]
        for sdr in sdrs:
            assert sdr["runway_no"] is not None

    def test_ndb_station_defaults(self, dualiang_xlsx_bytes: bytes) -> None:
        rows = _parse_station_sheet(dualiang_xlsx_bytes)
        ndbs = [r for r in rows if r["station_type"] == "NDB"]
        assert len(ndbs) > 0

    def test_vor_station_defaults(self, dualiang_xlsx_bytes: bytes) -> None:
        rows = _parse_station_sheet(dualiang_xlsx_bytes)
        vors = [r for r in rows if r["station_type"] == "VOR"]
        assert len(vors) > 0
        vor = vors[0]
        assert vor["reflection_net_hag"] is not None
        assert vor["reflection_diameter"] is not None
        assert vor["b_to_center_distance"] == pytest.approx(6.75)
        # C# 对齐：bAntennaH 默认 1.2（Excel 中 VOR 此列无值）
        assert vor["b_antenna_h"] == pytest.approx(1.2)
        assert vor["center_antenna_h"] == pytest.approx(1.2)

    def test_all_stations_have_required_fields(self, dualiang_xlsx_bytes: bytes) -> None:
        rows = _parse_station_sheet(dualiang_xlsx_bytes)
        for r in rows:
            assert r["station_type"]
            assert r["name"]
            assert "longitude" in r
            assert "latitude" in r

    def test_missing_sheet_raises(self, empty_airport_xlsx_bytes: bytes) -> None:
        with pytest.raises(AirportImportParseError):
            _parse_station_sheet(empty_airport_xlsx_bytes)


# --- Orchestrator Tests ---

class TestImportOrchestrator:
    def test_full_import_creates_airport_runways_stations(self, dualiang_xlsx_bytes: bytes) -> None:
        with _create_import_test_session() as session:
            result = _import_airport_from_excel(
                session=session, excel_bytes=dualiang_xlsx_bytes, original_filename="XX机场.xlsx",
            )
            session.commit()
            assert result["id"] is not None
            assert result["airportName"] == "XX机场"
            assert result["runwayCount"] == 4
            assert result["stationCount"] == 21
            airport = session.get(Airport, result["id"])
            assert airport is not None
            assert airport.name == "XX机场"
            runway_count = session.scalar(
                select(func.count()).select_from(Runway).where(Runway.airport_id == airport.id)
            )
            assert runway_count == 4
            station_count = session.scalar(
                select(func.count()).select_from(Station).where(Station.airport_id == airport.id)
            )
            assert station_count == 21

    def test_upsert_replaces_existing(self, dualiang_xlsx_bytes: bytes) -> None:
        with _create_import_test_session() as session:
            r1 = _import_airport_from_excel(session=session, excel_bytes=dualiang_xlsx_bytes, original_filename="XX机场.xlsx")
            session.commit()
            r2 = _import_airport_from_excel(session=session, excel_bytes=dualiang_xlsx_bytes, original_filename="XX机场.xlsx")
            session.commit()
            assert r2["id"] == r1["id"]
            assert r2["runwayCount"] == 4

    def test_rolls_back_on_invalid_excel(self) -> None:
        with _create_import_test_session() as session:
            with pytest.raises(AirportImportParseError):
                _import_airport_from_excel(session=session, excel_bytes=b"not an excel file", original_filename="bad.xlsx")
            airports = session.scalar(select(func.count()).select_from(Airport))
            assert airports == 0

    def test_airport_coordinates_from_first_runway(self, dualiang_xlsx_bytes: bytes) -> None:
        with _create_import_test_session() as session:
            result = _import_airport_from_excel(
                session=session, excel_bytes=dualiang_xlsx_bytes, original_filename="XX机场.xlsx",
            )
            session.commit()
            airport = session.get(Airport, result["id"])
            assert airport.longitude is not None
            assert airport.latitude is not None
            assert airport.altitude is not None

    def test_empty_runway_sheet_falls_back_to_stations(self) -> None:
        wb = Workbook()
        ws1 = wb.create_sheet("Sheet1", 0)
        all_rw_headers = [
            "跑道号码", "跑道真方位（°）", "跑道长度（米）", "跑道宽度（米）",
            "跑道入口标高（米）", "最大适航机型（H：航空器高度）",
            "跑道中心点经度", "跑道中心点纬度", "机场基准点标高（米）",
            "跑道类型", "编码A", "编码B", "仪表着陆系统类别",
            "最大可起降航空器类别类别",
        ]
        for i, h in enumerate(all_rw_headers, start=1):
            ws1.cell(row=1, column=i, value=h)
        ws2 = wb.create_sheet("Sheet2")
        ws2.cell(row=4, column=1, value="NDB")
        ws2.cell(row=4, column=2, value="NDB_TEST")
        ws2.cell(row=4, column=3, value=300)
        ws2.cell(row=4, column=4, value=103.5)
        ws2.cell(row=4, column=5, value=30.5)
        ws2.cell(row=4, column=6, value=500)
        buf = BytesIO()
        wb.save(buf)
        with _create_import_test_session() as session:
            result = _import_airport_from_excel(
                session=session, excel_bytes=buf.getvalue(), original_filename="Test.xlsx",
            )
            session.commit()
            airport = session.get(Airport, result["id"])
            assert airport.longitude is not None
            assert airport.latitude is not None
            assert result["runwayCount"] == 0
            assert result["stationCount"] == 1


# --- API Integration Tests ---

class TestImportAirportApi:
    def test_post_single_file_returns_batch_response(self, dualiang_xlsx_bytes: bytes) -> None:
        with _create_import_test_session() as session:
            app.dependency_overrides[get_db_session] = lambda: session
            try:
                client = TestClient(app)
                response = client.post(
                    "/data-management/import/airports",
                    files=[("excelFiles", ("XX机场.xlsx", dualiang_xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
                )
                assert response.status_code == 201
                data = response.json()
                assert data["totalFiles"] == 1
                assert data["importedCount"] == 1
                assert data["skippedCount"] == 0
                item = data["items"][0]
                assert item["status"] == "imported"
                assert item["fileName"] == "XX机场.xlsx"
                assert item["airportId"] is not None
                assert item["airportName"] == "XX机场"
                assert item["runwayCount"] == 4
                assert item["stationCount"] == 21
            finally:
                app.dependency_overrides.pop(get_db_session, None)

    def test_post_multiple_files_imports_all(self, dualiang_xlsx_bytes: bytes) -> None:
        with _create_import_test_session() as session:
            app.dependency_overrides[get_db_session] = lambda: session
            try:
                client = TestClient(app)
                response = client.post(
                    "/data-management/import/airports",
                    files=[
                        ("excelFiles", ("XX机场.xlsx", dualiang_xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
                        ("excelFiles", ("Copy机场.xlsx", dualiang_xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
                    ],
                )
                assert response.status_code == 201
                data = response.json()
                assert data["totalFiles"] == 2
                assert data["importedCount"] == 2
                assert data["skippedCount"] == 0
                assert len(data["items"]) == 2
                assert data["items"][0]["airportName"] == "XX机场"
                assert data["items"][1]["airportName"] == "Copy机场"
            finally:
                app.dependency_overrides.pop(get_db_session, None)

    def test_post_invalid_file_is_skipped(self, dualiang_xlsx_bytes: bytes) -> None:
        with _create_import_test_session() as session:
            app.dependency_overrides[get_db_session] = lambda: session
            try:
                client = TestClient(app)
                response = client.post(
                    "/data-management/import/airports",
                    files=[
                        ("excelFiles", ("XX机场.xlsx", dualiang_xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
                        ("excelFiles", ("bad.txt", b"not excel", "text/plain")),
                    ],
                )
                assert response.status_code == 201
                data = response.json()
                assert data["totalFiles"] == 2
                assert data["importedCount"] == 1
                assert data["skippedCount"] == 1
                assert data["items"][1]["status"] == "skipped"
                assert data["items"][1]["fileName"] == "bad.txt"
            finally:
                app.dependency_overrides.pop(get_db_session, None)
