from dataclasses import dataclass, field
from io import BytesIO
import re
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


EXPECTED_SHEET_NAME = "Sheet1"
EXPECTED_HEADERS = ["障碍物名称", "经度", "纬度", "顶部高程"]
_DMS_PATTERN = re.compile(r'^\s*(\d+)[°º]\s*(\d+)[\'′]\s*(\d+(?:\.\d+)?)["″]\s*$')


class PolygonObstacleExcelParseError(ValueError):
    pass


@dataclass(slots=True)
class PolygonObstaclePoint:
    row_number: int
    longitude_text: str
    latitude_text: str
    longitude_decimal: float
    latitude_decimal: float


@dataclass(slots=True)
class PolygonObstacle:
    name: str
    top_elevation: float
    points: list[PolygonObstaclePoint] = field(default_factory=list)


# 解析单个经纬度 DMS 文本为十进制度数。
def _parse_dms(value: str, *, field_name: str, row_number: int) -> float:
    match = _DMS_PATTERN.match(value)
    if match is None:
        raise PolygonObstacleExcelParseError(
            f"row {row_number} {field_name} has invalid DMS value: {value}"
        )

    degrees = float(match.group(1))
    minutes = float(match.group(2))
    seconds = float(match.group(3))
    return degrees + minutes / 60 + seconds / 3600


# 解析固定模板的障碍物 Excel 文件。
def parse_polygon_obstacle_excel(excel_bytes: bytes) -> list[PolygonObstacle]:
    try:
        workbook = load_workbook(filename=BytesIO(excel_bytes), data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise PolygonObstacleExcelParseError("invalid excel file") from exc

    if EXPECTED_SHEET_NAME not in workbook.sheetnames:
        raise PolygonObstacleExcelParseError(
            f"missing worksheet: {EXPECTED_SHEET_NAME}"
        )

    worksheet = workbook[EXPECTED_SHEET_NAME]
    headers = [worksheet.cell(row=1, column=index).value for index in range(1, 5)]
    if headers != EXPECTED_HEADERS:
        raise PolygonObstacleExcelParseError(
            f"invalid header row: expected {EXPECTED_HEADERS}, got {headers}"
        )

    obstacles_by_name: dict[str, PolygonObstacle] = {}
    obstacles: list[PolygonObstacle] = []

    for row_number in range(2, worksheet.max_row + 1):
        name = worksheet.cell(row=row_number, column=1).value
        longitude_text = worksheet.cell(row=row_number, column=2).value
        latitude_text = worksheet.cell(row=row_number, column=3).value
        top_elevation = worksheet.cell(row=row_number, column=4).value

        row_values = [name, longitude_text, latitude_text, top_elevation]
        if all(value in (None, "") for value in row_values):
            continue

        if not isinstance(name, str) or not name.strip():
            raise PolygonObstacleExcelParseError(
                f"row {row_number} 障碍物名称 is required"
            )
        if not isinstance(longitude_text, str) or not longitude_text.strip():
            raise PolygonObstacleExcelParseError(f"row {row_number} 经度 is required")
        if not isinstance(latitude_text, str) or not latitude_text.strip():
            raise PolygonObstacleExcelParseError(f"row {row_number} 纬度 is required")
        if top_elevation in (None, ""):
            raise PolygonObstacleExcelParseError(
                f"row {row_number} 顶部高程 is required"
            )

        try:
            top_elevation_float = float(top_elevation)
        except (TypeError, ValueError) as exc:
            raise PolygonObstacleExcelParseError(
                f"row {row_number} 顶部高程 must be numeric"
            ) from exc

        point = PolygonObstaclePoint(
            row_number=row_number,
            longitude_text=longitude_text.strip(),
            latitude_text=latitude_text.strip(),
            longitude_decimal=_parse_dms(
                longitude_text.strip(), field_name="经度", row_number=row_number
            ),
            latitude_decimal=_parse_dms(
                latitude_text.strip(), field_name="纬度", row_number=row_number
            ),
        )

        obstacle = obstacles_by_name.get(name)
        if obstacle is None:
            obstacle = PolygonObstacle(
                name=name.strip(), top_elevation=top_elevation_float
            )
            obstacles_by_name[obstacle.name] = obstacle
            obstacles.append(obstacle)
        elif obstacle.top_elevation != top_elevation_float:
            raise PolygonObstacleExcelParseError(
                f"row {row_number} 顶部高程 does not match previous rows for {obstacle.name}"
            )

        obstacle.points.append(point)

    if not obstacles:
        raise PolygonObstacleExcelParseError("worksheet contains no data rows")

    return obstacles
