from dataclasses import dataclass
from io import BytesIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.application.polygon_obstacle_excel_parser import (
    EXPECTED_SHEET_NAME,
    _detect_obstacle_template,
    _parse_dms,
    _parse_dms_components,
)


class PointObstacleExcelParseError(ValueError):
    pass


@dataclass(slots=True)
class PointObstacle:
    name: str
    row_number: int
    longitude_text: str
    latitude_text: str
    longitude_decimal: float
    latitude_decimal: float
    top_elevation: float


def parse_point_obstacle_excel(excel_bytes: bytes) -> list[PointObstacle]:
    try:
        workbook = load_workbook(filename=BytesIO(excel_bytes), data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise PointObstacleExcelParseError("invalid excel file") from exc

    if EXPECTED_SHEET_NAME not in workbook.sheetnames:
        raise PointObstacleExcelParseError(
            f"missing worksheet: {EXPECTED_SHEET_NAME}"
        )

    worksheet = workbook[EXPECTED_SHEET_NAME]
    template = _detect_obstacle_template(worksheet)

    obstacles_by_name: dict[str, PointObstacle] = {}
    obstacles: list[PointObstacle] = []

    for row_number in range(2, worksheet.max_row + 1):
        name = worksheet.cell(row=row_number, column=1).value

        if template == "8col":
            lon_deg = worksheet.cell(row=row_number, column=2).value
            lon_min = worksheet.cell(row=row_number, column=3).value
            lon_sec = worksheet.cell(row=row_number, column=4).value
            lat_deg = worksheet.cell(row=row_number, column=5).value
            lat_min = worksheet.cell(row=row_number, column=6).value
            lat_sec = worksheet.cell(row=row_number, column=7).value
            top_elevation = worksheet.cell(row=row_number, column=8).value
            longitude_text = ""
            latitude_text = ""
            row_values = [
                name, lon_deg, lon_min, lon_sec, lat_deg, lat_min, lat_sec,
                top_elevation,
            ]
        else:
            longitude_text = worksheet.cell(row=row_number, column=2).value
            latitude_text = worksheet.cell(row=row_number, column=3).value
            top_elevation = worksheet.cell(row=row_number, column=4).value
            lon_deg = lon_min = lon_sec = lat_deg = lat_min = lat_sec = None
            row_values = [name, longitude_text, latitude_text, top_elevation]

        if all(value in (None, "") for value in row_values):
            continue

        if not isinstance(name, str) or not name.strip():
            raise PointObstacleExcelParseError(
                f"row {row_number} 障碍物名称 is required"
            )

        if template == "8col":
            if lon_deg is None and lon_min is None and lon_sec is None:
                raise PointObstacleExcelParseError(
                    f"row {row_number} 经度 is required"
                )
            if lat_deg is None and lat_min is None and lat_sec is None:
                raise PointObstacleExcelParseError(
                    f"row {row_number} 纬度 is required"
                )
        else:
            if not isinstance(longitude_text, str) or not longitude_text.strip():
                raise PointObstacleExcelParseError(
                    f"row {row_number} 经度 is required"
                )
            if not isinstance(latitude_text, str) or not latitude_text.strip():
                raise PointObstacleExcelParseError(
                    f"row {row_number} 纬度 is required"
                )

        if top_elevation in (None, ""):
            raise PointObstacleExcelParseError(
                f"row {row_number} 顶部高程 is required"
            )

        try:
            top_elevation_float = float(top_elevation)
        except (TypeError, ValueError) as exc:
            raise PointObstacleExcelParseError(
                f"row {row_number} 顶部高程 must be numeric"
            ) from exc

        if template == "8col":
            longitude_decimal = _parse_dms_components(lon_deg, lon_min, lon_sec)
            latitude_decimal = _parse_dms_components(lat_deg, lat_min, lat_sec)
        else:
            longitude_decimal = _parse_dms(
                longitude_text.strip(), field_name="经度", row_number=row_number
            )
            latitude_decimal = _parse_dms(
                latitude_text.strip(), field_name="纬度", row_number=row_number
            )
            longitude_text = longitude_text.strip()
            latitude_text = latitude_text.strip()

        obstacle = PointObstacle(
            name=name.strip(),
            row_number=row_number,
            longitude_text=longitude_text,
            latitude_text=latitude_text,
            longitude_decimal=longitude_decimal,
            latitude_decimal=latitude_decimal,
            top_elevation=top_elevation_float,
        )

        existing_obstacle = obstacles_by_name.get(obstacle.name)
        if existing_obstacle is None:
            obstacles_by_name[obstacle.name] = obstacle
            obstacles.append(obstacle)
            continue

        if (
            existing_obstacle.longitude_decimal == obstacle.longitude_decimal
            and existing_obstacle.latitude_decimal == obstacle.latitude_decimal
            and existing_obstacle.top_elevation == obstacle.top_elevation
        ):
            continue

        raise PointObstacleExcelParseError(
            f"row {row_number} 同名障碍物的坐标或顶部高程不一致: {obstacle.name}"
        )

    if not obstacles:
        raise PointObstacleExcelParseError("worksheet contains no data rows")

    return obstacles
