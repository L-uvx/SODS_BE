import math
import os
import re
from decimal import Decimal
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select, delete
from sqlalchemy.orm import Session

from app.models.airport import Airport
from app.models.runway import Runway
from app.models.station import Station


_DMS_DELIMITER_RE = re.compile(
    r"[度分秒°′″'’\"\\\":EN＇＂〞º\u00b4ʹ\u02da]"
)

_NUMBER_RE = re.compile(r"\d+(?:\.\d+)?")

_KM_RE = re.compile(r"[kK][mM]|千米|公里")
_NM_RE = re.compile(r"[nN][mM]|海里")

_RUNWAY_NO_RE = re.compile(r"(\d{1,2}[LRC]?)")


def _parse_degree(value: Any) -> float | None:
    """
    解析经纬度值，兼容十进制度与度分秒格式。
    """
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError(f"cannot parse boolean as degree: {value}")
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None
    parts = _DMS_DELIMITER_RE.split(text)
    if len(parts) == 1:
        return float(parts[0].strip())

    numeric_parts = [p.strip() for p in parts if p.strip()]
    if len(numeric_parts) == 1:
        return float(numeric_parts[0])

    d = float(numeric_parts[0])
    m = float(numeric_parts[1]) if len(numeric_parts) > 1 else 0.0
    s = float(numeric_parts[2]) if len(numeric_parts) > 2 else 0.0

    if m < 0 or m >= 60:
        raise ValueError(f"minutes out of range: {m}")
    if s < 0 or s >= 60:
        raise ValueError(f"seconds out of range: {s}")

    return d + m / 60.0 + s / 3600.0


def _get_number_from_string(value: Any) -> float | None:
    """
    从字符串中提取数值，支持 km/nm 单位转换。
    """
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError(f"cannot parse boolean as number: {value}")
    if isinstance(value, (int, float)):
        return float(value)

    cleaned = str(value).replace(" ", "")
    match = _NUMBER_RE.search(cleaned)
    if match is None:
        # 对齐 C# GetNumberFromString：无数字时返回原字符串
        try:
            return float(cleaned)
        except ValueError:
            return None

    num = float(match.group())

    suffix = cleaned[: match.start()] + cleaned[match.end() :]
    if _KM_RE.search(suffix):
        return num * 1000.0
    if _NM_RE.search(suffix):
        return num * 1852.0

    return num


def _extract_runway_info(name: Any) -> str | None:
    """
    从台站名称中提取跑道编号。
    """
    if name is None:
        return None
    if isinstance(name, bool):
        return None
    text = str(name).strip()
    if not text:
        return None

    match = _RUNWAY_NO_RE.search(text)
    if match:
        return match.group()

    parts = text.split()
    return parts[-1] if parts else None


def _int_floor(value: float) -> float:
    """
    保留两位小数，向下取整。
    """
    return float(math.floor(value * 100) / 100)


class AirportImportParseError(ValueError):
    pass


def _open_workbook(excel_bytes: bytes):
    try:
        return load_workbook(filename=BytesIO(excel_bytes), data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise AirportImportParseError("invalid excel file") from exc


def _safe_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


_MAXIMUM_AIRWORTHINESS_MAP: dict[str, float] = {
    '车辆（H≤6米）': 0,
    '车辆(H≤6米)': 0,
    '中型航空器(6米≤H≤14米)': 1,
    '大型航空器(14米≤H≤20米)': 2,
    '特大型航空器(20米≤H≤25米)': 3,
}


def _map_maximum_airworthiness(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError(f"cannot parse boolean as number: {value}")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    result = _MAXIMUM_AIRWORTHINESS_MAP.get(text)
    if result is not None:
        return result
    return _get_number_from_string(value)


_RUNWAY_ENUM_MAPS: dict[str, dict[str | int | float, str]] = {
    'runway_type': {
        '非仪表跑道': '非仪表跑道', 0: '非仪表跑道',
        '非精密进近跑道': '非精密进近跑道', 1: '非精密进近跑道',
        '精密进近跑道': '精密进近跑道', 2: '精密进近跑道',
    },
    'station_sub_type': {
        'I类': 'I', 0: 'I',
        'II类': 'II', 1: 'II',
        'III类': 'III', 2: 'III',
    },
    'maximum_type_aircraft': {
        'D类和D类以上': 'D类和D类以上', 0: 'D类和D类以上',
        'C类和C类以下': 'C类和C类以下', 1: 'C类和C类以下',
        'B类和B类以下': 'B类和B类以下', 2: 'B类和B类以下',
    },
}

_ANTENNA_UNIT_NUMBER_MAP: dict[str | int | float, str] = {
    '小孔径（11单元及以下）': '10',
    'S': '10', 0: '10',
    '中孔径（12至15单元）': '14',
    'M': '14', 1: '14',
    '大孔径（16单元及以上）': '20',
    'L': '20', 2: '20',
}


def _map_enum_field(value: Any, mapping: dict[str | int | float, str]) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text in mapping:
        return mapping[text]
    try:
        num = float(text)
        if num in mapping:
            return mapping[num]
    except ValueError:
        pass
    return text


def _map_antenna_unit_number(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError(f"cannot parse boolean: {value}")
    if isinstance(value, (int, float)):
        result = _ANTENNA_UNIT_NUMBER_MAP.get(int(value))
        if result is not None:
            return result
        return str(int(value)) if isinstance(value, (int, float)) else str(value)
    text = str(value).strip()
    if not text:
        return None
    result = _ANTENNA_UNIT_NUMBER_MAP.get(text)
    if result is not None:
        return result
    try:
        num = float(text)
        int_result = _ANTENNA_UNIT_NUMBER_MAP.get(int(num))
        if int_result is not None:
            return int_result
    except ValueError:
        pass
    return text


def _parse_runway_sheet(excel_bytes: bytes) -> list[dict[str, Any]]:
    workbook = _open_workbook(excel_bytes)
    if len(workbook.worksheets) == 0:
        raise AirportImportParseError("工作簿中没有工作表")
    ws = workbook.worksheets[0]

    rows: list[dict[str, Any]] = []
    row_index = 2
    while True:
        cell_a = ws.cell(row=row_index, column=1).value
        if cell_a is None:
            break
        try:
            runway_number = str(cell_a).strip()
            enter_height = ws.cell(row=row_index, column=5).value
            if enter_height is not None:
                enter_height = float(str(enter_height).replace(" ", ""))
            rows.append({
                'run_number': runway_number,
                'name': runway_number,
                'direction': _get_number_from_string(ws.cell(row=row_index, column=2).value),
                'length': _get_number_from_string(ws.cell(row=row_index, column=3).value),
                'width': _get_number_from_string(ws.cell(row=row_index, column=4).value),
                'enter_height': enter_height,
                'maximum_airworthiness': _map_maximum_airworthiness(ws.cell(row=row_index, column=6).value),
                'longitude': _parse_degree(ws.cell(row=row_index, column=7).value),
                'latitude': _parse_degree(ws.cell(row=row_index, column=8).value),
                'altitude': _get_number_from_string(ws.cell(row=row_index, column=9).value),
                'runway_type': _map_enum_field(ws.cell(row=row_index, column=10).value, _RUNWAY_ENUM_MAPS['runway_type']),
                'runway_code_a': _safe_str(ws.cell(row=row_index, column=11).value),
                'runway_code_b': _safe_str(ws.cell(row=row_index, column=12).value),
                'station_sub_type': _map_enum_field(ws.cell(row=row_index, column=13).value, _RUNWAY_ENUM_MAPS['station_sub_type']),
                'maximum_type_aircraft': _map_enum_field(ws.cell(row=row_index, column=14).value, _RUNWAY_ENUM_MAPS['maximum_type_aircraft']),
            })
        except AirportImportParseError:
            raise
        except Exception as exc:
            raise AirportImportParseError(f"跑道数据解析失败（行数：{row_index}") from exc
        row_index += 1

    return rows


def _parse_station_sheet(excel_bytes: bytes) -> list[dict[str, Any]]:
    workbook = _open_workbook(excel_bytes)
    if len(workbook.worksheets) < 2:
        raise AirportImportParseError("缺少台站工作表（第2个Sheet）")
    ws = workbook.worksheets[1]

    rows: list[dict[str, Any]] = []
    row_index = 4
    max_row = ws.max_row or 0

    while True:
        try:
            cell_a = ws.cell(row=row_index, column=1).value
            if cell_a is None and row_index > max_row:
                break
            if cell_a is None:
                row_index += 1
                continue

            station_type = str(cell_a).strip()
            if station_type not in _STATION_PARSERS:
                row_index += 1
                continue

            base = _read_station_base_fields(ws, row_index)
            parsed = _STATION_PARSERS[station_type](ws, row_index, base)
            if parsed is not None:
                rows.append(parsed)
        except AirportImportParseError:
            raise
        except Exception as exc:
            raise AirportImportParseError(
                f"台站数据解析失败请检查表格格式或度分秒符号是否正确（行数：{row_index}）"
            ) from exc
        row_index += 1

    return rows


def _read_station_base_fields(ws, row_index: int) -> dict[str, Any]:
    return {
        'station_type': str(ws.cell(row=row_index, column=1).value or '').strip(),
        'name': str(ws.cell(row=row_index, column=2).value or '').strip(),
        'frequency': _get_number_from_string(ws.cell(row=row_index, column=3).value),
        'longitude': _parse_degree(ws.cell(row=row_index, column=4).value),
        'latitude': _parse_degree(ws.cell(row=row_index, column=5).value),
        'altitude': _read_altitude(ws, row_index),
        'coverage_radius_raw': _get_number_from_string(ws.cell(row=row_index, column=7).value),
        'antenna_hag_raw': _get_number_from_string(ws.cell(row=row_index, column=8).value),
        'unit_number': _map_antenna_unit_number(ws.cell(row=row_index, column=9).value),
    }


def _read_altitude(ws, row_index: int) -> float | None:
    return _get_number_from_string(ws.cell(row=row_index, column=6).value)


# --- Per-type parsers ---

def _parse_loc(ws, row_index, base):
    ah = base['antenna_hag_raw']
    if ah is None or ah == 0:
        ah = 3.0
    un = base['unit_number']
    if not un:
        raise AirportImportParseError(f"LOC天线单元个数不能为空（行数：{row_index}）")
    der = _get_number_from_string(ws.cell(row=row_index, column=10).value)
    if der is None:
        raise AirportImportParseError(f"LOC与跑道末端距离不能为空（行数：{row_index}）")
    rn = _safe_str(ws.cell(row=row_index, column=18).value)
    if rn:
        rn = rn.replace(' ', '')
    alt = base.get('altitude') or 0.0
    return {
        **base, 'station_type': 'LOC', 'antenna_hag': ah, 'coverage_radius': 46300.0,
        'fly_height': _int_floor((alt + 600) * 100 / 100), 'unit_number': un,
        'distance_endo_runway': der, 'runway_no': rn, 'station_sub_type': None,
    }


def _parse_gp(ws, row_index, base):
    if base['frequency'] is None:
        raise AirportImportParseError(f"GP频率不能为空（行数：{row_index}）")
    ah = base['antenna_hag_raw']
    # C# 对齐：GP 天线高度单元格非空但解析失败时抛异常
    raw_ah_cell = ws.cell(row=row_index, column=8).value
    if raw_ah_cell is not None and ah is None:
        raise AirportImportParseError(f"GP天线离地高解析失败（行数：{row_index}）")
    if ah is not None and 3 < ah < 6:
        ah *= 3
    elif ah is not None and 6 < ah < 10:
        ah *= 1.5
    elif ah is None or ah == 0:
        ah = 12.9
    dtr = _get_number_from_string(ws.cell(row=row_index, column=11).value)
    if dtr is None:
        raise AirportImportParseError(f"GP后撤距离不能为空（行数：{row_index}）")
    dvr = _get_number_from_string(ws.cell(row=row_index, column=12).value)
    if dvr is None:
        raise AirportImportParseError(f"GP距离跑道中线距离不能为空（行数：{row_index}）")
    da = _get_number_from_string(ws.cell(row=row_index, column=13).value)
    if da is None:
        da = 3.0
    ga = _get_number_from_string(ws.cell(row=row_index, column=14).value)
    rn = _safe_str(ws.cell(row=row_index, column=18).value)
    if rn:
        rn = rn.replace(' ', '')
    alt = base.get('altitude') or 0.0
    return {
        **base, 'station_type': 'GP', 'antenna_hag': ah, 'coverage_radius': 18520.0,
        'downward_angle': da, 'fly_height': _int_floor((alt + 600) * 100 / 100),
        'distance_to_runway': dtr, 'distance_v_to_runway': dvr, 'antenna_height': ga,
        'runway_no': rn, 'station_sub_type': None,
    }


def _parse_vor(ws, row_index, base):
    rnh = _get_number_from_string(ws.cell(row=row_index, column=15).value)
    if rnh is None:
        rnh = 25.0
    rd = _get_number_from_string(ws.cell(row=row_index, column=16).value)
    if rd is None:
        rd = 30.0
    bh = _get_number_from_string(ws.cell(row=row_index, column=17).value)
    if bh is not None and bh <= 2:
        bah = bh
        cah = bh
    else:
        bah = 1.2
        cah = 1.2
    cr = base['coverage_radius_raw']
    if cr is None:
        cr = 37040.0
    alt = base.get('altitude') or 0.0
    return {
        **base, 'station_type': 'VOR', 'reflection_net_hag': rnh, 'reflection_diameter': rd,
        'b_antenna_h': bah, 'center_antenna_h': cah, 'b_to_center_distance': 6.75,
        'coverage_radius': cr, 'fly_height': _int_floor((alt + 600) * 100 / 100),
    }


def _parse_ndb(ws, row_index, base):
    cr = base['coverage_radius_raw']
    if cr is None:
        cr = 37040.0
    alt = base.get('altitude') or 0.0
    return {
        **base, 'station_type': 'NDB', 'antenna_hag': base['antenna_hag_raw'] or 0.0,
        'coverage_radius': cr, 'fly_height': _int_floor((alt + 600) * 100 / 100),
    }


def _parse_mb(ws, row_index, base):
    ah = base['antenna_hag_raw'] or 0.0
    rn = _safe_str(ws.cell(row=row_index, column=18).value)
    if rn:
        rn = rn.replace(' ', '')
    alt = base.get('altitude') or 0.0
    return {
        **base, 'station_type': 'MB', 'antenna_hag': ah, 'coverage_radius': 30.0,
        'fly_height': _int_floor((alt + ah + 40) * 100 / 100), 'runway_no': rn,
    }


def _parse_radar(ws, row_index, base):
    alt = base.get('altitude') or 0.0
    return {
        **base, 'station_type': 'RADAR', 'antenna_hag': base['antenna_hag_raw'] or 0.0,
        'coverage_radius': 30000.0, 'fly_height': _int_floor((alt + 600) * 100 / 100),
    }


def _parse_surface_detection_radar(ws, row_index, base):
    rn = _safe_str(ws.cell(row=row_index, column=18).value)
    if rn:
        rn = rn.replace(' ', '')
    alt = base.get('altitude') or 0.0
    return {
        **base, 'station_type': 'Surface_Detection_Radar',
        'antenna_hag': base['antenna_hag_raw'] or 0.0, 'coverage_radius': 30000.0,
        'fly_height': _int_floor((alt + 600) * 100 / 100), 'runway_no': rn,
    }


def _make_40_type_parser(station_type: str):
    def _parse(ws, row_index, base):
        cr = base['coverage_radius_raw']
        if cr is None:
            cr = 37040.0
        ah = base['antenna_hag_raw'] or 0.0
        alt = base.get('altitude') or 0.0
        return {
            **base, 'station_type': station_type, 'antenna_hag': ah,
            'coverage_radius': cr, 'fly_height': _int_floor((alt + ah + 40) * 100 / 100),
        }
    return _parse


_STATION_PARSERS: dict[str, Any] = {
    'LOC': _parse_loc,
    'GP': _parse_gp,
    'VOR': _parse_vor,
    'NDB': _parse_ndb,
    'MB': _parse_mb,
    'PSR/SSR': _parse_radar,
    'RADAR': _parse_radar,
    'Radar': _parse_radar,
    'SMR': _parse_surface_detection_radar,
    'Surface_Detection_Radar': _parse_surface_detection_radar,
    'ADS-B': _make_40_type_parser('ADS_B'),
    'ADS_B': _make_40_type_parser('ADS_B'),
    'VHF': _make_40_type_parser('VHF'),
    'HF': _make_40_type_parser('HF'),
    'WeatherRadar': _make_40_type_parser('WeatherRadar'),
    'WindRadar': _make_40_type_parser('WindRadar'),
    'GBAS': _make_40_type_parser('GBAS'),
}


def _import_airport_from_excel(
    *, session: Session, excel_bytes: bytes, original_filename: str,
) -> dict[str, Any]:
    airport_name = os.path.splitext(os.path.basename(original_filename))[0]
    if not airport_name:
        raise AirportImportParseError("文件名不能为空")

    runway_rows = _parse_runway_sheet(excel_bytes)
    station_rows = _parse_station_sheet(excel_bytes)

    existing = _find_airport_by_name(session, airport_name)
    if existing is not None:
        _delete_airport_children(session, existing.id)
        airport = existing
        airport.name = airport_name
    else:
        airport = Airport(name=airport_name)

    if runway_rows:
        first = runway_rows[0]
        airport.longitude = _to_decimal(first['longitude'])
        airport.latitude = _to_decimal(first['latitude'])
        airport.altitude = _to_decimal(first['altitude'])

    session.add(airport)
    session.flush()
    airport_id = airport.id

    runway_map: dict[str, Runway] = {}
    for rw_data in runway_rows:
        runway = Runway(
            airport_id=airport_id,
            run_number=rw_data['run_number'],
            name=rw_data['name'],
            direction=_to_decimal(rw_data.get('direction')),
            length=_to_decimal(rw_data.get('length')),
            width=_to_decimal(rw_data.get('width')),
            enter_height=_to_decimal(rw_data.get('enter_height')),
            maximum_airworthiness=_to_decimal(rw_data.get('maximum_airworthiness')),
            longitude=_to_decimal(rw_data.get('longitude')),
            latitude=_to_decimal(rw_data.get('latitude')),
            altitude=_to_decimal(rw_data.get('altitude')),
            runway_type=rw_data.get('runway_type'),
            runway_code_a=rw_data.get('runway_code_a'),
            runway_code_b=rw_data.get('runway_code_b'),
            station_sub_type=rw_data.get('station_sub_type'),
            maximum_type_aircraft=rw_data.get('maximum_type_aircraft'),
        )
        session.add(runway)
        session.flush()
        runway_map[rw_data['run_number']] = runway

    for st_data in station_rows:
        _resolve_station_runway(st_data, runway_map)
        _fill_station_altitude(st_data, runway_map, airport.altitude)
        station = Station(
            airport_id=airport_id,
            station_type=st_data.get('station_type'),
            name=st_data.get('name'),
            frequency=_to_decimal(st_data.get('frequency')),
            longitude=_to_decimal(st_data.get('longitude')),
            latitude=_to_decimal(st_data.get('latitude')),
            altitude=_to_decimal(st_data.get('altitude')),
            coverage_radius=_to_decimal(st_data.get('coverage_radius')),
            fly_height=_to_decimal(st_data.get('fly_height')),
            antenna_hag=_to_decimal(st_data.get('antenna_hag')),
            runway_no=st_data.get('runway_no'),
            reflection_net_hag=_to_decimal(st_data.get('reflection_net_hag')),
            center_antenna_h=_to_decimal(st_data.get('center_antenna_h')),
            b_antenna_h=_to_decimal(st_data.get('b_antenna_h')),
            b_to_center_distance=_to_decimal(st_data.get('b_to_center_distance')),
            reflection_diameter=_to_decimal(st_data.get('reflection_diameter')),
            downward_angle=_to_decimal(st_data.get('downward_angle')),
            distance_to_runway=_to_decimal(st_data.get('distance_to_runway')),
            distance_v_to_runway=_to_decimal(st_data.get('distance_v_to_runway')),
            distance_endo_runway=_to_decimal(st_data.get('distance_endo_runway')),
            unit_number=st_data.get('unit_number'),
            antenna_height=_to_decimal(st_data.get('antenna_height')),
            station_sub_type=st_data.get('station_sub_type'),
        )
        session.add(station)

    if not runway_rows and station_rows:
        loc_stations = [s for s in station_rows if s['station_type'] == 'LOC']
        ref = loc_stations[0] if loc_stations else station_rows[0]
        airport.longitude = _to_decimal(ref['longitude'])
        airport.latitude = _to_decimal(ref['latitude'])
        airport.altitude = _to_decimal(ref['altitude'])

    session.flush()

    return {
        'id': airport_id, 'airportName': airport_name,
        'runwayCount': len(runway_rows), 'stationCount': len(station_rows), 'warnings': [],
    }


def _resolve_station_runway(st_data: dict[str, Any], runway_map: dict[str, Runway]) -> None:
    st_type = st_data.get('station_type', '')
    if st_type == 'Surface_Detection_Radar':
        rn = st_data.get('runway_no')
        if rn and rn in runway_map:
            st_data['station_sub_type'] = runway_map[rn].station_sub_type
        return
    if st_type in ('LOC', 'GP', 'MB'):
        rn = st_data.get('runway_no') or _extract_runway_info(st_data.get('name'))
        if rn and rn in runway_map:
            runway = runway_map[rn]
            st_data['runway_no'] = rn
            st_data['station_sub_type'] = runway.station_sub_type
        elif st_type in ('LOC', 'GP'):
            raise AirportImportParseError(f"台站 {st_data.get('name')} 无法匹配跑道（{rn}）")
        else:
            st_data['runway_no'] = rn


def _fill_station_altitude(
    st_data: dict[str, Any],
    runway_map: dict[str, Runway],
    airport_altitude: Any,
) -> None:
    if st_data.get('altitude') is not None:
        return

    rn = st_data.get('runway_no')
    if rn and rn in runway_map and runway_map[rn].altitude is not None:
        st_data['altitude'] = float(runway_map[rn].altitude)

    if st_data.get('altitude') is None and airport_altitude is not None:
        st_data['altitude'] = float(airport_altitude)

    alt = st_data.get('altitude')
    if alt is None:
        return

    st_type = st_data.get('station_type', '')
    if st_type in ('LOC', 'NDB', 'VOR', 'GP', 'RADAR', 'Surface_Detection_Radar'):
        st_data['fly_height'] = _int_floor((alt + 600) * 100 / 100)
    else:
        ah = st_data.get('antenna_hag', 0.0) or 0.0
        st_data['fly_height'] = _int_floor((alt + ah + 40) * 100 / 100)


def _find_airport_by_name(session: Session, name: str) -> Airport | None:
    return session.scalar(select(Airport).where(Airport.name == name))


def _delete_airport_children(session: Session, airport_id: int) -> None:
    session.execute(delete(Station).where(Station.airport_id == airport_id))
    session.execute(delete(Runway).where(Runway.airport_id == airport_id))
    session.flush()


def _to_decimal(value: Any) -> Any:
    if value is None:
        return None
    return Decimal(str(value))
